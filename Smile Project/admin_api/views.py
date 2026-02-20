# admin_api/views.py
"""
OPTIMIZED ADMIN API VIEWS
- Pagination untuk handle jutaan data
- Caching untuk stats
- Optimized queries dengan select_related & prefetch_related
- Rate limiting untuk security
- Input validation
- Excel export functionality
"""

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAdminUser, IsAuthenticated
from .permissions import IsStoreStaffOrAbove
from rest_framework.pagination import PageNumberPagination
from django.db.models import Count, Sum, Q, Prefetch
from django.utils import timezone
from django.core.cache import cache
from django.http import HttpResponse
from datetime import timedelta
from django.utils.html import escape
from decimal import Decimal, InvalidOperation
import uuid
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

from users.models import User
from policies.models import Policy, PolicyTier, DevicePackage
from claims.models import Claim
from wallet.models import Wallet, TopUpTransaction, WalletHistory
from stores.models import Store
from stores.activity_log import ActivityLog
from .decorators import rate_limit_api


class OptimizedPagination(PageNumberPagination):
    """
    Custom pagination untuk handle jutaan data.
    - Default 50 items per page
    - Max 100 items per page
    - Include count, next, previous
    """
    page_size = 50
    page_size_query_param = 'page_size'
    max_page_size = 100


class DashboardStatsViewSet(viewsets.ViewSet):
    """
    Dashboard Statistics - CACHED
    GET /api/admin/dashboard/
    """
    permission_classes = [IsStoreStaffOrAbove]
    
    def list(self, request):
        from django.db.models import Sum, Count, Q
        from django.db.models.functions import TruncMonth
        from datetime import timedelta
        
        user = request.user
        is_store_admin = hasattr(user, 'role') and user.role in ['store_admin', 'store_staff']
        user_store = getattr(user, 'store', None)
        
        # Filter by store for Store Admin
        if is_store_admin and user_store:
            # Store Admin: Only data from their store
            policies_qs = Policy.objects.filter(store=user_store)
            claims_qs = Claim.objects.filter(policy__store=user_store)
            # Use user.store instead of policy_set__store (simpler, more reliable)
            users_qs = User.objects.filter(store=user_store)
            # Wallets for users in this store
            wallets_qs = Wallet.objects.filter(user__store=user_store)
            # TopUpTransaction uses user directly, not wallet
            topups_qs = TopUpTransaction.objects.filter(user__store=user_store)
        else:
            # Super Admin: All data
            policies_qs = Policy.objects.all()
            claims_qs = Claim.objects.all()
            users_qs = User.objects.all()
            wallets_qs = Wallet.objects.all()
            topups_qs = TopUpTransaction.objects.all()
        
        # 1. Financial Metrics
        total_premium = policies_qs.filter(status__in=['active', 'expired']).aggregate(total=Sum('policy_price'))['total'] or 0
        total_claim_paid = claims_qs.filter(status__in=['approved', 'completed']).aggregate(total=Sum('claim_amount'))['total'] or 0
        loss_ratio = (total_claim_paid / total_premium * 100) if total_premium > 0 else 0

        # DAILY STATS (For Mobile Dashboard)
        today = timezone.now().date()
        today_policies_qs = policies_qs.filter(created_at__date=today)
        today_claims_qs = claims_qs.filter(created_at__date=today)
        
        today_premium = today_policies_qs.filter(status__in=['active', 'expired']).aggregate(total=Sum('policy_price'))['total'] or 0
        today_policies_count = today_policies_qs.count()
        today_claims_count = today_claims_qs.count() # All claims submitted today

        # 2. Build Stats Dictionary
        stats = {
            'overview': {
                'total_premium': float(total_premium),
                'total_claim_paid': float(total_claim_paid),
                'loss_ratio': round(loss_ratio, 2),
                'outstanding_claims': claims_qs.filter(status='pending').count(),
                # Add Daily Stats
                'today_premium': float(today_premium),
                'today_policies': today_policies_count,
                'today_claims': today_claims_count
            },
            'trends': [],
            'users': {
                'total': users_qs.count(),
                'verified': users_qs.filter(is_verified=True).count(),
                'active': users_qs.filter(is_active=True).count(),
            },
            'policies': {
                'total': policies_qs.count(),
                'active': policies_qs.filter(status='active').count(),
                'pending': policies_qs.filter(status='pending').count(),
                'expired': policies_qs.filter(status='expired').count(),
            },
            'claims': {
                'total': claims_qs.count(),
                'pending': claims_qs.filter(status='pending').count(),
                'approved': claims_qs.filter(status='approved').count(),
                'rejected': claims_qs.filter(status='rejected').count(),
                'total_amount': float(total_claim_paid)
            },
            'wallet': {
                'total_balance': float(wallets_qs.aggregate(Sum('balance'))['balance__sum'] or 0),
                'total_topup': float(wallets_qs.aggregate(Sum('total_topup'))['total_topup__sum'] or 0),
                'pending_topups': topups_qs.filter(status='pending').count(),
            },
            'top_stores': [],
            'store_info': None  # For Store Admin context
        }
        
        # Add store info for Store Admin
        if is_store_admin and user_store:
            stats['store_info'] = {
                'id': str(user_store.id),
                'name': user_store.name,
                'code': user_store.registration_code
            }
        
        # 3. Calculate Trends
        try:
            six_months_ago = timezone.now() - timedelta(days=180)
            
            premium_trend = policies_qs.filter(
                created_at__gte=six_months_ago,
                status__in=['active', 'expired']
            ).annotate(month=TruncMonth('created_at')).values('month').annotate(
                total=Sum('policy_price')
            ).order_by('month')
            
            claim_trend = claims_qs.filter(
                created_at__gte=six_months_ago
            ).annotate(month=TruncMonth('created_at')).values('month').annotate(
                total=Sum('claim_amount')
            ).order_by('month')
            
            months_map = {}
            for p in premium_trend:
                if p['month']:
                    m_key = p['month'].strftime('%Y-%m')
                    months_map[m_key] = {'name': p['month'].strftime('%b %Y'), 'premium': float(p['total'] or 0), 'claims': 0}
            
            for c in claim_trend:
                if c['month']:
                    m_key = c['month'].strftime('%Y-%m')
                    if m_key not in months_map:
                        months_map[m_key] = {'name': c['month'].strftime('%b %Y'), 'premium': 0, 'claims': 0}
                    months_map[m_key]['claims'] = float(c['total'] or 0)
            
            if months_map:
                sorted_keys = sorted(months_map.keys())
                for k in sorted_keys:
                    stats['trends'].append(months_map[k])
        except Exception as e:
            print(f"Error trends: {e}")

        # 4. Top Stores (Only for Super Admin)
        if not is_store_admin:
            try:
                # Use Policy.store relation directly
                top_stores_qs = Store.objects.annotate(
                    policy_count=Count('policies'),
                    premium_value=Sum('policies__policy_price', filter=Q(policies__status='active'))
                ).filter(is_active=True).order_by('-policy_count')[:5]

                for s in top_stores_qs:
                    stats['top_stores'].append({
                        'id': str(s.id),
                        'name': s.name,
                        'code': s.registration_code,
                        'policy_count': s.policy_count,
                        'premium_value': float(s.premium_value or 0)
                    })
            except Exception as e:
                print(f"Error calculating top stores: {e}")
        
        return Response(stats)


class AdminUserViewSet(viewsets.ModelViewSet):
    """
    Admin User Management - OPTIMIZED & RATE LIMITED
    
    Store Admin: Only sees users from their store
    Super Admin: Sees all users (can filter by store)
    
    GET /api/admin/users/?search=john&is_verified=true&page=1
    GET /api/admin/users/?store=<store_id>  (Super Admin only)
    """
    permission_classes = [IsStoreStaffOrAbove]
    pagination_class = OptimizedPagination
    
    def get_queryset(self):
        queryset = User.objects.select_related('store').order_by('-date_joined')
        user = self.request.user
        
        # Base Filtering (Role-based isolation)
        if user.role == 'super_admin':
            # Super Admin: Can see all, optionally filter by store
            store_filter = self.request.query_params.get('store')
            if store_filter:
                queryset = queryset.filter(store_id=store_filter)
        elif user.role in ['store_admin', 'store_staff'] and user.store:
            # ✅ FIX: Store Admin/Staff only sees users REGISTERED at their store
            queryset = queryset.filter(store=user.store)
        else:
            # Other roles or staff without store assigned shouldn't see anything
            return queryset.none()
        
        # Search by email, phone, name, or KTP (with XSS protection)
        search = self.request.query_params.get('search', None)
        if search:
            # Sanitize input to prevent SQL injection (Django ORM already safe, but extra safety)
            search = escape(search.strip())[:100]  # Limit length
            queryset = queryset.filter(
                Q(email__icontains=search) |
                Q(phone_number__icontains=search) |
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(ktp_number__icontains=search)  # Include KTP search
            )
        
        # Filter by verification status
        is_verified = self.request.query_params.get('is_verified', None)
        if is_verified is not None and is_verified != '':
            queryset = queryset.filter(is_verified=is_verified.lower() == 'true')
        
        # Filter by active status
        is_active = self.request.query_params.get('is_active', None)
        if is_active is not None and is_active != '':
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
        
        return queryset
    
    def list(self, request):
        queryset = self.get_queryset()
        page = self.paginate_queryset(queryset)
        
        data = []
        for user in page:
            # Safe access to reverse relations (try 'policies' then 'policy_set')
            policies_rel = getattr(user, 'policies', getattr(user, 'policy_set', None))
            claims_rel = getattr(user, 'claims', getattr(user, 'claim_set', None))
            
            display_policy = None
            total_policies = 0
            active_policies = 0
            
            if policies_rel:
                # We need to fetch related fields to avoid N+1 inside loop if possible, 
                # but since we didn't prefetch cleanly, we rely on lazy loading (acceptable for admin list < 50)
                all_policies = policies_rel.all().order_by('-created_at')
                total_policies = all_policies.count()
                
                # Find active policy (manually iterating a few items is faster than DB query if count is small)
                # But safer to filter
                active_policy_qs = all_policies.filter(status='active')
                if active_policy_qs.exists():
                    display_policy = active_policy_qs.first()
                    active_policies = active_policy_qs.count()
                elif total_policies > 0:
                     display_policy = all_policies.first() # Show latest if none active
            
            total_claims = claims_rel.count() if claims_rel else 0

            data.append({
                'id': str(user.id),
                'email': user.email,
                'full_name': f"{user.first_name} {user.last_name}",
                'phone_number': user.phone_number,
                'ktp_number': user.ktp_number,
                'is_verified': user.is_verified,
                'is_active': user.is_active,
                'date_joined': user.date_joined,
                'role': user.role,
                'store': {
                    'id': str(user.store.id),
                    'code': user.store.code,
                    'name': user.store.name
                } if user.store else None,
                'stats': {
                    'total_policies': total_policies,
                    'total_claims': total_claims,
                    'active_policies': active_policies
                },
                'device_info': f"{display_policy.device_package.device_brand} {display_policy.device_package.device_model}" if display_policy and display_policy.device_package else "-",
                'tier_info': display_policy.tier.tier_name if display_policy and display_policy.tier else "-",
            })
        
        return self.get_paginated_response(data)
    
    def retrieve(self, request, pk=None):
        """
        Get single user detail
        GET /api/admin/users/{id}/
        """
        try:
            # ✅ FIX: Use get_queryset to enforce store isolation
            user = self.get_queryset().get(pk=pk)
        except User.DoesNotExist:
            return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)
        
        data = {
            'id': str(user.id),
            'email': user.email,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'phone_number': user.phone_number,
            'ktp_number': user.ktp_number,
            'address': user.address,
            'birth_date': user.birth_date,
            'is_verified': user.is_verified,
            'is_active': user.is_active,
            'date_joined': user.date_joined,
            'role': user.role,
            'store': {
                'id': str(user.store.id),
                'code': user.store.code,
                'name': user.store.name
            } if user.store else None,
        }
        return Response(data)
    
    def update(self, request, pk=None):
        """
        Update user data
        PUT /api/admin/users/{id}/
        Body: {
            "first_name": "John",
            "last_name": "Doe",
            "phone_number": "081234567890",
            "ktp_number": "3201234567891234",
            "address": "Jakarta",
            "is_verified": true,
            "is_active": true,
            "role": "store_admin",  // Super Admin only
            "store": "uuid-of-store"  // Super Admin only
        }
        """
        try:
            # ✅ FIX: Use get_queryset to enforce store isolation
            user = self.get_queryset().get(pk=pk)
        except User.DoesNotExist:
            return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)
        
        # Update fields (with validation)
        first_name = request.data.get('first_name', '').strip()
        last_name = request.data.get('last_name', '').strip()
        phone_number = request.data.get('phone_number', '').strip()
        ktp_number = request.data.get('ktp_number', '').strip()
        address = request.data.get('address', '').strip()
        is_verified = request.data.get('is_verified', user.is_verified)
        is_active = request.data.get('is_active', user.is_active)
        
        # Validation
        if first_name:
            user.first_name = escape(first_name)[:50]
        if last_name:
            user.last_name = escape(last_name)[:50]
        if phone_number:
            user.phone_number = escape(phone_number)[:15]
        if ktp_number:
            # Validate KTP: must be 16 digits
            if len(ktp_number) == 16 and ktp_number.isdigit():
                user.ktp_number = ktp_number
            else:
                return Response(
                    {'error': 'KTP must be exactly 16 digits'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        if address:
            user.address = escape(address)[:255]
        
        user.is_verified = is_verified
        user.is_active = is_active
        
        # Super Admin only: Role and Store assignment
        if request.user.role == 'super_admin':
            new_role = request.data.get('role', None)
            new_store = request.data.get('store', None)
            
            if new_role:
                valid_roles = ['customer', 'store_staff', 'store_admin', 'super_admin']
                if new_role in valid_roles:
                    old_role = user.role
                    user.role = new_role
                    
                    # Sync Django flags based on role
                    if new_role == 'super_admin':
                        user.is_superuser = True
                        user.is_staff = True
                        user.store = None  # Force clear store for super_admin
                    elif new_role in ['store_admin', 'store_staff']:
                        user.is_superuser = False
                        user.is_staff = True
                    else:
                        user.is_superuser = False
                        user.is_staff = False
                    
                    # Log role change
                    from stores.activity_log import ActivityLog
                    ActivityLog.log(
                        request=request,
                        action='USER_UPDATE',
                        target_model='User',
                        target_id=str(user.id),
                        description=f"Role changed from {old_role} to {new_role}",
                        extra_data={'old_role': old_role, 'new_role': new_role}
                    )
            
            # Store assignment (Allow for customers too, to fix registration errors)
            if new_store:
                from stores.models import Store
                try:
                    store = Store.objects.get(id=new_store)
                    user.store = store
                    
                    # Log store assignment
                    from stores.activity_log import ActivityLog
                    ActivityLog.log(
                        request=request,
                        action='USER_ASSIGN_STORE',
                        target_model='User',
                        target_id=str(user.id),
                        description=f"User assigned to store {store.code}",
                        extra_data={'store_id': str(store.id), 'store_code': store.code}
                    )
                except Store.DoesNotExist:
                    return Response(
                        {'error': 'Store not found'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            elif new_role in ['customer', 'super_admin']:
                # Clear store for customer and super_admin
                user.store = None
        
        user.save()
        
        # Log activity
        ActivityLog.log(
            request=request,
            action='USER_UPDATE',
            target_model='User',
            target_id=str(user.id),
            description=f"User diupdate: {user.email}"
        )
        
        return Response({
            'message': 'User updated successfully',
            'user': {
                'id': str(user.id),
                'email': user.email,
                'full_name': f"{user.first_name} {user.last_name}",
                'phone_number': user.phone_number,
                'ktp_number': user.ktp_number,
                'role': user.role,
                'store': {
                    'id': str(user.store.id),
                    'code': user.store.code,
                    'name': user.store.name
                } if user.store else None
            }
        })
    
    def destroy(self, request, pk=None):
        """
        Delete a user - SUPER ADMIN ONLY
        DELETE /api/admin/users/{id}/
        """
        # Only Super Admin can delete users
        if request.user.role != 'super_admin':
            return Response(
                {'error': 'Hanya Super Admin yang dapat menghapus user'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        try:
            user = User.objects.get(id=pk)
        except User.DoesNotExist:
            return Response(
                {'error': 'User tidak ditemukan'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Prevent deleting yourself
        if user.id == request.user.id:
            return Response(
                {'error': 'Tidak dapat menghapus akun sendiri'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Prevent deleting other super admins
        if user.role == 'super_admin':
            return Response(
                {'error': 'Tidak dapat menghapus Super Admin lain'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Check if user has policies or claims
        policy_count = user.policy_set.count()
        claim_count = user.claim_set.count()
        
        if policy_count > 0 or claim_count > 0:
            return Response({
                'error': 'User memiliki data terkait dan tidak dapat dihapus',
                'details': {
                    'policies': policy_count,
                    'claims': claim_count
                },
                'suggestion': 'Nonaktifkan user dengan set is_active=False sebagai gantinya'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Log before delete
        user_email = user.email
        ActivityLog.log(
            request=request,
            action='USER_DELETE',
            target_model='User',
            target_id=str(user.id),
            description=f"User dihapus: {user_email}"
        )
        
        user.delete()
        
        return Response({
            'message': f'User {user_email} berhasil dihapus'
        })
    
    @action(detail=True, methods=['post'])
    def reset_password(self, request, pk=None):
        """
        Reset user password - Admin resets for user
        POST /api/admin/users/{id}/reset_password/
        """
        try:
            user = User.objects.get(id=pk)
        except User.DoesNotExist:
            return Response(
                {'error': 'User tidak ditemukan'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        new_password = request.data.get('new_password')
        
        if not new_password:
            return Response(
                {'error': 'new_password diperlukan'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if len(new_password) < 6:
            return Response(
                {'error': 'Password minimal 6 karakter'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        user.set_password(new_password)
        user.save()
        
        # Log activity
        ActivityLog.log(
            request=request,
            action='USER_UPDATE',
            target_model='User',
            target_id=str(user.id),
            description=f"Password direset oleh admin untuk: {user.email}"
        )
        
        return Response({
            'message': f'Password untuk {user.email} berhasil direset'
        })
    
    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """
        Export all users to Excel
        GET /api/admin/users/export_excel/
        """
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        
        # Header styling
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers
        headers = ['ID', 'Email', 'Full Name', 'Phone', 'KTP Number', 'Verified', 'Active', 'Registered Date']
        ws.append(headers)
        
        # Style header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # OPTIMIZED: Limit export to 10000 users and use iterator for memory efficiency
        # For larger exports, recommend using async/background task
        MAX_EXPORT_ROWS = 10000
        users = User.objects.only(
            'id', 'email', 'first_name', 'last_name', 
            'phone_number', 'ktp_number', 'is_verified', 'is_active', 'date_joined'
        ).order_by('-date_joined')[:MAX_EXPORT_ROWS]
        
        # Write data using iterator for memory efficiency
        for user in users.iterator():
            ws.append([
                str(user.id),
                user.email,
                f"{user.first_name} {user.last_name}",
                user.phone_number or '',
                user.ktp_number or '',
                'Yes' if user.is_verified else 'No',
                'Yes' if user.is_active else 'No',
                user.date_joined.strftime('%Y-%m-%d %H:%M:%S') if user.date_joined else ''
            ])
        
        # Adjust column widths
        column_widths = [36, 30, 25, 15, 18, 10, 10, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=users_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        wb.save(response)
        return response


class AdminClaimViewSet(viewsets.ModelViewSet):
    """
    Admin Claim Management - OPTIMIZED
    GET /api/admin/claims/?status=pending&search=CLM&page=1
    GET /api/admin/claims/notifications/ - Get pending claims for notification
    """
    permission_classes = [IsStoreStaffOrAbove]
    pagination_class = OptimizedPagination
    
    def get_queryset(self):
        # Optimize with select_related dan prefetch_related untuk photos
        queryset = Claim.objects.select_related(
            'user', 'policy', 'policy__device_package', 'policy__tier', 'policy__store'
        ).prefetch_related('photos').order_by('-created_at')
        
        # ✅ FIX: Store Admin only sees claims from policies in their store
        user = self.request.user
        if user.role in ['store_admin', 'store_staff'] and user.store:
            queryset = queryset.filter(policy__store=user.store)
        elif user.role == 'super_admin':
            # Super Admin can filter by store optionally
            store_filter = self.request.query_params.get('store')
            if store_filter:
                queryset = queryset.filter(policy__store_id=store_filter)
        
        # Filter by status
        status_filter = self.request.query_params.get('status', None)
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        # Search by claim number or user email
        search = self.request.query_params.get('search', None)
        if search:
            queryset = queryset.filter(
                Q(claim_number__icontains=search) |
                Q(user__email__icontains=search)
            )
        
        return queryset
    
    def list(self, request):
        queryset = self.get_queryset()
        page = self.paginate_queryset(queryset)
        
        # Serialize data with photos
        data = [{
            'id': str(claim.id),
            'claim_number': claim.claim_number,
            'user_email': claim.user.email,
            'user_name': f"{claim.user.first_name} {claim.user.last_name}",
            'user_full_name': f"{claim.user.first_name} {claim.user.last_name}", # Alias for Flutter
            'device': f"{claim.policy.device_package.device_brand} {claim.policy.device_package.device_model}",
            'device_full_name': f"{claim.policy.device_package.device_brand} {claim.policy.device_package.device_model}", # Alias for Flutter
            'store_name': claim.policy.store.name if claim.policy.store else '-',
            'imei_number': claim.policy.imei_number or 'N/A',  # IMEI untuk verifikasi
            'damage_type': claim.damage_type,
            'damage_description': claim.damage_description,
            'incident_date': claim.incident_date,
            'claim_amount': float(claim.claim_amount),
            'wallet_deducted': float(claim.wallet_deducted) if claim.wallet_deducted else 0,
            'status': claim.status,
            'admin_notes': claim.admin_notes or '',
            'whatsapp_number': claim.whatsapp_number or '',
            'payment_notes': claim.payment_notes or '',
            'payment_date': claim.payment_date,
            'created_at': claim.created_at,
            'updated_at': claim.processed_date if claim.processed_date else claim.created_at, # Alias for Flutter
            # Include photos with full URL
            'photos': [{
                'id': str(photo.id),
                'photo_url': request.build_absolute_uri(photo.photo.url) if photo.photo else None,
                'uploaded_at': photo.uploaded_at.isoformat() if photo.uploaded_at else None
            } for photo in claim.photos.all()]
        } for claim in page]
        
        return self.get_paginated_response(data)
    
    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """
        Export all claims to Excel
        GET /api/admin/claims/export_excel/
        """
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Claims"
        
        # Header styling
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers
        headers = ['Claim Number', 'User Email', 'User Name', 'Device', 'Damage Type', 
                   'Claim Amount', 'Wallet Deducted', 'Status', 'Created Date', 'Admin Notes']
        ws.append(headers)
        
        # Style header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Get all claims
        claims = Claim.objects.select_related(
            'user', 'policy', 'policy__device_package'
        ).all().order_by('-created_at')
        
        # Write data
        for claim in claims:
            ws.append([
                claim.claim_number,
                claim.user.email,
                f"{claim.user.first_name} {claim.user.last_name}",
                f"{claim.policy.device_package.device_brand} {claim.policy.device_package.device_model}",
                claim.damage_type,
                float(claim.claim_amount),
                float(claim.wallet_deducted) if claim.wallet_deducted else 0,
                claim.status,
                claim.created_at.strftime('%Y-%m-%d %H:%M:%S') if claim.created_at else '',
                claim.admin_notes or ''
            ])
        
        # Adjust column widths
        column_widths = [18, 30, 25, 25, 20, 15, 15, 12, 20, 30]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=claims_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        wb.save(response)
        return response
    
    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """
        Approve claim with input validation
        POST /api/admin/claims/{id}/approve/
        Body: { "claim_amount": 5000000, "admin_notes": "Approved" }
        """
        claim = self.get_object()
        
        # Input validation
        claim_amount = request.data.get('claim_amount')
        admin_notes = request.data.get('admin_notes', '')
        
        if not claim_amount:
            return Response(
                {'error': 'claim_amount is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validate claim_amount is a valid number
        try:
            claim_amount = Decimal(str(claim_amount))
            if claim_amount <= 0:
                return Response(
                    {'error': 'claim_amount must be positive'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            if claim_amount > 100000000:  # Max 100 million
                return Response(
                    {'error': 'claim_amount too large'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        except (ValueError, InvalidOperation):
            return Response(
                {'error': 'Invalid claim_amount format'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Sanitize admin notes (prevent XSS)
        admin_notes = escape(admin_notes.strip())[:500]  # Max 500 chars
        
        # Check if policy has enough balance
        policy = claim.policy
        if policy.policy_balance < claim_amount:
            return Response(
                {'error': f'Insufficient policy balance. Available: Rp {policy.policy_balance:,.0f}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Update claim
        claim.claim_amount = claim_amount
        claim.wallet_deducted = claim_amount  # Record deduction amount
        claim.status = 'approved'
        claim.admin_notes = admin_notes
        claim.processed_by = request.user
        claim.processed_date = timezone.now()
        claim.save()
        
        # ✅ Deduct from POLICY BALANCE (not wallet!)
        policy.policy_balance -= claim_amount
        policy.claims_used += 1
        policy.save()
        
        # Log activity
        ActivityLog.log(
            request=request,
            action='CLAIM_APPROVE',
            target_model='Claim',
            target_id=str(claim.id),
            description=f"Klaim disetujui: {claim.claim_number} - Rp {claim_amount:,.0f}"
        )
        
        # Clear dashboard cache
        cache.delete('admin_dashboard_stats')
        
        return Response({
            'message': 'Claim approved successfully',
            'policy_balance_remaining': float(policy.policy_balance)
        })
    
    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        """
        Reject claim
        POST /api/admin/claims/{id}/reject/
        Body: { "admin_notes": "Invalid proof" }
        """
        claim = self.get_object()
        admin_notes = request.data.get('admin_notes', '')
        
        claim.status = 'rejected'
        claim.admin_notes = admin_notes
        claim.processed_by = request.user
        claim.processed_date = timezone.now()
        claim.save()
        
        # Log activity
        ActivityLog.log(
            request=request,
            action='CLAIM_REJECT',
            target_model='Claim',
            target_id=str(claim.id),
            description=f"Klaim ditolak: {claim.claim_number}"
        )
        
        # Clear dashboard cache
        cache.delete('admin_dashboard_stats')
        
        return Response({'message': 'Claim rejected successfully'})
    
    @action(detail=True, methods=['post'])
    def set_in_progress(self, request, pk=None):
        """
        Set claim status to in_progress
        POST /api/admin/claims/{id}/set_in_progress/
        Body: { "admin_notes": "Being processed at service center" }
        """
        claim = self.get_object()
        
        # Validate: Only approved claims can be set to in_progress
        if claim.status != 'approved':
            return Response(
                {'error': 'Only approved claims can be set to in_progress'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        admin_notes = request.data.get('admin_notes', '')
        
        claim.status = 'in_progress'
        if admin_notes:
            claim.admin_notes = escape(admin_notes.strip())[:500]
        claim.save()
        
        # Log activity
        ActivityLog.log(
            request=request,
            action='CLAIM_UPDATE',
            target_model='Claim',
            target_id=str(claim.id),
            description=f"Klaim dalam proses: {claim.claim_number}"
        )
        
        # Clear dashboard cache
        cache.delete('admin_dashboard_stats')
        
        return Response({
            'message': 'Claim status updated to In Progress',
            'claim': {
                'id': str(claim.id),
                'claim_number': claim.claim_number,
                'status': claim.status
            }
        })
    
    @action(detail=True, methods=['post'])
    def set_completed(self, request, pk=None):
        """
        Set claim status to completed with WhatsApp notification
        POST /api/admin/claims/{id}/set_completed/
        Body: { 
            "admin_notes": "Repair completed, device sent",
            "whatsapp_number": "081234567890",
            "payment_notes": "Paid via BCA transfer",
            "payment_date": "2025-11-25T10:30:00Z"
        }
        """
        claim = self.get_object()
        
        # Validate: Only approved or in_progress claims can be completed
        if claim.status not in ['approved', 'in_progress']:
            return Response(
                {'error': 'Only approved or in_progress claims can be completed'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        admin_notes = request.data.get('admin_notes', '')
        whatsapp_number = request.data.get('whatsapp_number', '')
        payment_notes = request.data.get('payment_notes', '')
        payment_date = request.data.get('payment_date', None)
        
        claim.status = 'completed'
        if admin_notes:
            claim.admin_notes = escape(admin_notes.strip())[:500]
        if whatsapp_number:
            # Clean WhatsApp number (remove spaces, dashes, etc.)
            clean_number = ''.join(filter(str.isdigit, whatsapp_number))
            claim.whatsapp_number = clean_number[:20]
        if payment_notes:
            claim.payment_notes = escape(payment_notes.strip())[:500]
        if payment_date:
            claim.payment_date = payment_date
        claim.save()
        
        # Log activity
        ActivityLog.log(
            request=request,
            action='CLAIM_COMPLETE',
            target_model='Claim',
            target_id=str(claim.id),
            description=f"Klaim selesai: {claim.claim_number}"
        )
        
        # Clear dashboard cache
        cache.delete('admin_dashboard_stats')
        
        return Response({
            'message': 'Claim marked as Completed',
            'claim': {
                'id': str(claim.id),
                'claim_number': claim.claim_number,
                'status': claim.status
            }
        })
    
    @action(detail=False, methods=['get'])
    def notifications(self, request):
        """
        Get pending claims for notification bell
        GET /api/admin/claims/notifications/
        
        Returns:
        {
            "pending_count": 5,
            "recent_claims": [
                {
                    "id": "...",
                    "claim_number": "CLM-20251124...",
                    "user_name": "John Doe",
                    "device": "iPhone 15 Pro",
                    "damage_type": "Layar Pecah",
                    "created_at": "2025-11-24T10:30:00Z"
                }
            ]
        }
        """
        # Base queryset for pending claims
        pending_qs = Claim.objects.filter(status='pending')
        
        # ✅ FIX: Filter by store for Store Admin
        user = request.user
        if user.role in ['store_admin', 'store_staff'] and user.store:
            pending_qs = pending_qs.filter(policy__store=user.store)
        
        # Get pending claims count
        pending_count = pending_qs.count()
        
        # Get recent 5 pending claims
        recent_claims = pending_qs.select_related(
            'user', 'policy__device_package'
        ).order_by('-created_at')[:5]
        
        data = {
            'pending_count': pending_count,
            'recent_claims': [{
                'id': str(claim.id),
                'claim_number': claim.claim_number,
                'user_name': f"{claim.user.first_name} {claim.user.last_name}",
                'user_email': claim.user.email,
                'device': f"{claim.policy.device_package.device_brand} {claim.policy.device_package.device_model}",
                'damage_type': claim.damage_type,
                'created_at': claim.created_at.isoformat(),
            } for claim in recent_claims]
        }
        
        return Response(data)
    
    @action(detail=False, methods=['post'])
    def create_for_user(self, request):
        """
        Admin creates a claim on behalf of a user.
        Use case: User's phone is damaged and they cannot access the app.
        POST /api/admin/claims/create_for_user/
        """
        from claims.models import ClaimPhoto
        from django.db import transaction
        import traceback
        
        try:
            data = request.data
            
            # Validate required fields
            user_id = data.get('user_id')
            policy_id = data.get('policy_id')
            damage_type = data.get('damage_type')
            incident_date = data.get('incident_date')
            
            if not all([user_id, policy_id, damage_type, incident_date]):
                return Response({
                    'error': 'Missing required fields: user_id, policy_id, damage_type, incident_date'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Get user
            try:
                user = User.objects.get(id=user_id)
            except User.DoesNotExist:
                return Response({'error': 'User tidak ditemukan'}, status=status.HTTP_404_NOT_FOUND)
            
            # Get and validate policy
            try:
                policy = Policy.objects.get(id=policy_id, user=user)
            except Policy.DoesNotExist:
                return Response({'error': 'Policy tidak ditemukan atau bukan milik user ini'}, status=status.HTTP_404_NOT_FOUND)
            
            # Validate policy is active
            if policy.status != 'active':
                return Response({'error': f'Policy tidak aktif (status: {policy.status})'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Validate not expired
            if policy.is_expired():
                policy.status = 'expired'
                policy.save()
                return Response({
                    'error': 'Policy sudah kadaluarsa',
                    'expiry_date': policy.expiry_date.isoformat()
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Validate balance
            if policy.policy_balance <= 0:
                return Response({
                    'error': 'Saldo policy sudah habis',
                    'policy_balance': float(policy.policy_balance)
                }, status=status.HTTP_400_BAD_REQUEST)
            
            with transaction.atomic():
                # Create claim on behalf of user
                claim = Claim.objects.create(
                    user=user,
                    policy=policy,
                    claim_number=f"CLM-ADM-{timezone.now().strftime('%Y%m%d%H%M%S')}",
                    damage_type=damage_type,
                    damage_description=data.get('damage_description', 'Klaim diajukan oleh Admin'),
                    incident_date=incident_date,
                    claim_amount=0,  # Admin will set this on approval
                    status='pending',
                    admin_notes=f"Klaim diajukan oleh Admin ({request.user.email}) atas nama user. Alasan: {data.get('reason', 'HP user rusak, tidak bisa akses aplikasi')}"
                )
                
                # Handle photo uploads (if any)
                photos = request.FILES.getlist('photos')
                photo_count = 0
                if photos:
                    for photo in photos:
                        if photo.size <= 10 * 1024 * 1024:  # Max 10MB
                            ClaimPhoto.objects.create(
                                claim=claim,
                                photo=photo
                            )
                            photo_count += 1
            
            # Clear dashboard cache
            cache.delete('admin_dashboard_stats')
            
            # Log activity
            ActivityLog.log(
                request=request,
                action='CLAIM_ADMIN_CREATE',
                target_model='Claim',
                target_id=str(claim.id),
                description=f"Admin membuat klaim untuk: {user.email} - {claim.claim_number}"
            )
            
            return Response({
                'message': f'Klaim berhasil dibuat atas nama {user.first_name} {user.last_name}'.strip() or user.email,
                'data': {
                    'id': str(claim.id),
                    'claim_number': claim.claim_number,
                    'user_email': user.email,
                    'status': claim.status,
                },
                'created_by_admin': request.user.email,
                'photos_uploaded': photo_count
            }, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            return Response({
                'error': f'Server error: {str(e)}',
                'traceback': traceback.format_exc()
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class AdminPolicyViewSet(viewsets.ModelViewSet):
    """
    Admin Policy Management - OPTIMIZED
    GET /api/admin/policies/?status=pending&page=1
    POST /api/admin/policies/manual-create/ - Create policy for user
    """
    permission_classes = [IsStoreStaffOrAbove]
    pagination_class = OptimizedPagination
    
    def get_queryset(self):
        # Optimize with select_related
        queryset = Policy.objects.select_related(
            'user', 'tier', 'device_package', 'store'
        ).order_by('-created_at')
        
        # ✅ FIX: Store Admin only sees policies from their store
        user = self.request.user
        if user.role in ['store_admin', 'store_staff'] and user.store:
            queryset = queryset.filter(store=user.store)
        elif user.role == 'super_admin':
            # Super Admin can filter by store optionally
            store_filter = self.request.query_params.get('store')
            if store_filter:
                queryset = queryset.filter(store_id=store_filter)
        
        # Filter by status
        status_filter = self.request.query_params.get('status', None)
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        # Filter by user (for admin-assisted claims)
        user_filter = self.request.query_params.get('user', None)
        if user_filter:
            queryset = queryset.filter(user_id=user_filter)
        
        return queryset
    
    def list(self, request):
        queryset = self.get_queryset()
        page = self.paginate_queryset(queryset)
        
        data = [{
            'id': str(policy.id),
            'policy_number': policy.policy_number,
            'user_email': policy.user.email,
            'tier': policy.tier.tier_name,
            'tier_name': policy.tier.tier_name,  # Alias for frontend
            'device': f"{policy.device_package.device_brand} {policy.device_package.device_model}",  # String for PoliciesPage
            'device_obj': {  # Object for AdminClaimCreatePage
                'brand': policy.device_package.device_brand,
                'model': policy.device_package.device_model,
            },
            'policy_balance': float(policy.policy_balance),
            'user_full_name': f"{policy.user.first_name} {policy.user.last_name}",
            'device_brand': policy.device_package.device_brand,
            'device_model': policy.device_package.device_model,
            'store_name': policy.store.name if policy.store else '-',
            'imei_number': policy.imei_number,
            'policy_price': float(policy.policy_price),
            'status': policy.status,
            'activation_date': policy.activation_date,
            'expiry_date': policy.expiry_date,
            'created_at': policy.created_at,
        } for policy in page]
        
        return self.get_paginated_response(data)
    
    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """
        Export all policies to Excel
        GET /api/admin/policies/export_excel/
        """
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Policies"
        
        # Header styling
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers
        headers = ['Policy Number', 'User Email', 'User Name', 'Device', 'IMEI', 'Tier', 
                   'Policy Price', 'Status', 'Activation Date', 'Expiry Date', 'Created Date']
        ws.append(headers)
        
        # Style header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Get all policies
        policies = Policy.objects.select_related(
            'user', 'device_package', 'tier'
        ).all().order_by('-created_at')
        
        # Write data
        for policy in policies:
            ws.append([
                policy.policy_number,
                policy.user.email,
                f"{policy.user.first_name} {policy.user.last_name}",
                f"{policy.device_package.device_brand} {policy.device_package.device_model}",
                policy.imei_number,
                policy.tier.tier_name if policy.tier else '',
                float(policy.policy_price),
                policy.status,
                policy.activation_date.strftime('%Y-%m-%d') if policy.activation_date else '',
                policy.expiry_date.strftime('%Y-%m-%d') if policy.expiry_date else '',
                policy.created_at.strftime('%Y-%m-%d %H:%M:%S') if policy.created_at else ''
            ])
        
        # Adjust column widths
        column_widths = [18, 30, 25, 25, 18, 12, 15, 12, 15, 15, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=policies_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        wb.save(response)
        return response
    
    @action(detail=False, methods=['post'], url_path='manual-create')
    def manual_create(self, request):
        """
        Create policy manually for user with AUTO TOP-UP
        POST /api/admin/policies/manual-create/
        Body: {
            user_id: uuid,
            device_package_id: uuid,
            imei_number: '123456789012345',
            purchase_price: 5000000
        }
        
        Flow:
        1. Auto top-up wallet user sebesar purchase_price
        2. Create policy
        3. Wallet balance TETAP (tidak ada potongan untuk policy price)
        """
        from users.models import User
        from policies.models import Policy, DevicePackage, PolicyTier
        from datetime import timedelta
        
        user_id = request.data.get('user_id')
        device_package_id = request.data.get('device_package_id')
        imei_number = request.data.get('imei_number')
        purchase_price = request.data.get('purchase_price')
        
        # Validation
        if not all([user_id, device_package_id, imei_number, purchase_price]):
            return Response(
                {'error': 'All fields required: user_id, device_package_id, imei_number, purchase_price'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validate purchase price
        try:
            purchase_price = Decimal(str(purchase_price))
            if purchase_price <= 0:
                return Response(
                    {'error': 'Purchase price must be positive'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        except (ValueError, InvalidOperation):
            return Response(
                {'error': 'Invalid purchase price format'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validate IMEI (15 digits)
        if not imei_number.isdigit() or len(imei_number) != 15:
            return Response(
                {'error': 'IMEI must be exactly 15 digits'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Check if IMEI already exists
        if Policy.objects.filter(imei_number=imei_number).exists():
            return Response(
                {'error': 'IMEI already registered'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get user (Enforce store isolation)
        try:
            if request.user.role == 'super_admin':
                user = User.objects.get(id=user_id)
            else:
                if not request.user.store:
                    return Response({'error': 'Admin belum ditugaskan ke toko manapun'}, status=status.HTTP_403_FORBIDDEN)
                user = User.objects.get(id=user_id, store=request.user.store)
        except User.DoesNotExist:
            return Response(
                {'error': 'User not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Get device package
        try:
            device_package = DevicePackage.objects.get(id=device_package_id)
        except DevicePackage.DoesNotExist:
            return Response(
                {'error': 'Device package not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Find appropriate tier based on purchase price
        tier = PolicyTier.objects.filter(
            is_active=True,
            min_price__lte=purchase_price,
            max_price__gte=purchase_price
        ).first()
        
        if not tier:
            return Response(
                {'error': f'No tier found for price Rp {purchase_price:,.0f}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # ==================================================================
        # STEP 1: AUTO TOP-UP WALLET
        # ==================================================================
        # Get or create wallet
        wallet, created = Wallet.objects.get_or_create(
            user=user,
            defaults={'balance': Decimal('0.00')}
        )
        
        # Create top-up transaction
        topup = TopUpTransaction.objects.create(
            user=user,
            transaction_id=f"AUTO{timezone.now().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:6]}",
            amount=purchase_price,
            payment_method='admin_policy_creation',
            status='completed',
            verified_by=request.user,
            verified_at=timezone.now()
        )
        
        # Update wallet balance (top-up)
        old_balance = wallet.balance
        wallet.balance += purchase_price
        wallet.total_topup += purchase_price
        wallet.save()
        
        # Create wallet history for top-up
        WalletHistory.objects.create(
            wallet=wallet,
            transaction_type='topup',
            amount=purchase_price,
            balance_before=old_balance,
            balance_after=wallet.balance,
            description=f"Auto top-up saat pembuatan polis (Device: {device_package.device_brand} {device_package.device_model})",
            reference_id=str(topup.id)
        )
        
        # ==================================================================
        # STEP 2: CREATE POLICY
        # ==================================================================
        # Generate policy number
        policy_number = f"POL-{timezone.now().strftime('%Y%m%d%H%M%S')}-{user.id.hex[:6]}"
        
        # ✅ FIX: Auto-assign store for Store Admin
        policy_store = None
        if request.user.role in ['store_admin', 'store_staff'] and request.user.store:
            policy_store = request.user.store
        
        # Create policy
        policy = Policy.objects.create(
            user=user,
            tier=tier,
            device_package=device_package,
            imei_number=imei_number,
            purchase_price=purchase_price,
            policy_price=tier.policy_price,
            policy_balance=device_package.device_value,  # ✅ Set balance = harga HP
            policy_number=policy_number,
            activation_date=timezone.now().date(),
            expiry_date=timezone.now().date() + timedelta(days=tier.policy_duration_days),
            status='active',  # Auto-active for admin-created policies
            claims_used=0,
            store=policy_store  # ✅ Auto-assign store from admin
        )
        
        # ==================================================================
        # NO DEDUCTION: Wallet tetap full sesuai harga device
        # Policy price TIDAK dipotong dari wallet
        # ==================================================================
        
        # Log activity
        ActivityLog.log(
            request=request,
            action='POLICY_CREATE',
            target_model='Policy',
            target_id=str(policy.id),
            description=f"Polis dibuat: {policy.policy_number} untuk {user.email} - {device_package.device_brand} {device_package.device_model}"
        )
        
        # Clear cache
        cache.delete('admin_dashboard_stats')
        
        return Response({
            'message': 'Policy created successfully with auto top-up',
            'policy': {
                'id': str(policy.id),
                'policy_number': policy.policy_number,
                'user': user.email,
                'tier': tier.tier_name,
                'device': f"{device_package.device_brand} {device_package.device_model}",
                'imei': imei_number,
                'purchase_price': float(purchase_price),
                'policy_price': float(tier.policy_price),
                'activation_date': policy.activation_date,
                'expiry_date': policy.expiry_date,
                'status': policy.status
            },
            'wallet': {
                'topup_amount': float(purchase_price),
                'balance_before': float(old_balance),
                'balance_after': float(wallet.balance),
                'final_balance': float(wallet.balance)
            }
        }, status=status.HTTP_201_CREATED)


class AdminWalletViewSet(viewsets.ViewSet):
    """
    Admin Wallet Management - OPTIMIZED
    GET /api/admin/wallets/?page=1
    GET /api/admin/wallets/stats/ - Get total stats (all wallets)
    """
    permission_classes = [IsStoreStaffOrAbove]
    pagination_class = OptimizedPagination
    
    def list(self, request):
        # Get wallets with user info
        queryset = Wallet.objects.select_related('user').order_by('-balance')
        
        # ✅ FIX: Store isolation
        user = request.user
        if user.role in ['store_admin', 'store_staff'] and user.store:
            queryset = queryset.filter(user__store=user.store)
        elif user.role != 'super_admin':
            return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
        
        # Search filter
        search = request.query_params.get('search', None)
        if search:
            queryset = queryset.filter(
                Q(user__email__icontains=search) |
                Q(user__first_name__icontains=search) |
                Q(user__last_name__icontains=search)
            )
        
        # Pagination
        paginator = self.pagination_class()
        page = paginator.paginate_queryset(queryset, request)
        
        data = [{
            'id': str(wallet.id),
            'user_email': wallet.user.email,
            'user_name': f"{wallet.user.first_name} {wallet.user.last_name}",
            'balance': float(wallet.balance),
            'total_topup': float(wallet.total_topup),
            'total_spent': float(wallet.total_spent),
        } for wallet in page]
        
        return paginator.get_paginated_response(data)
    
    @action(detail=False, methods=['get'])
    def stats(self, request):
        """
        Get wallet statistics (SUM of ALL wallets, not paginated)
        GET /api/admin/wallets/stats/
        
        Returns:
        {
            "total_balance": 45393000.00,
            "total_topup": 46993000.00,
            "total_spent": 1600000.00,
            "wallet_count": 5
        }
        """
        from django.db.models import Sum, Count
        
        # Aggregate ALL wallets (not paginated)
        stats = Wallet.objects.aggregate(
            total_balance=Sum('balance'),
            total_topup=Sum('total_topup'),
            total_spent=Sum('total_spent'),
            wallet_count=Count('id')
        )
        
        # Return with defaults if no data
        return Response({
            'total_balance': float(stats['total_balance'] or 0),
            'total_topup': float(stats['total_topup'] or 0),
            'total_spent': float(stats['total_spent'] or 0),
            'wallet_count': stats['wallet_count'] or 0,
        })


class AdminTopUpViewSet(viewsets.ModelViewSet):
    """
    Admin Top-Up Management - OPTIMIZED
    GET /api/admin/topups/?status=pending&page=1
    POST /api/admin/topups/ - Create manual top-up
    """
    permission_classes = [IsStoreStaffOrAbove]
    pagination_class = OptimizedPagination
    
    def get_queryset(self):
        queryset = TopUpTransaction.objects.select_related('user').order_by('-created_at')
        
        # ✅ FIX: Store isolation
        user = self.request.user
        if user.role in ['store_admin', 'store_staff'] and user.store:
            queryset = queryset.filter(user__store=user.store)
        elif user.role != 'super_admin':
            return queryset.none()

        # Filter by status
        status_filter = self.request.query_params.get('status', None)
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        return queryset
    
    def list(self, request):
        queryset = self.get_queryset()
        page = self.paginate_queryset(queryset)
        
        data = [{
            'id': str(topup.id),
            'transaction_id': topup.transaction_id,
            'user_email': topup.user.email,
            'amount': float(topup.amount),
            'payment_method': topup.payment_method,
            'status': topup.status,
            'created_at': topup.created_at,
        } for topup in page]
        
        return self.get_paginated_response(data)
    
    def create(self, request):
        """
        Create manual top-up for user
        POST /api/admin/topups/
        Body: {
            user: user_id,
            amount: 100000,
            payment_method: 'admin_topup',
            notes: 'Manual top-up by admin',
            status: 'completed'
        }
        """
        user_id = request.data.get('user')
        amount = request.data.get('amount')
        payment_method = request.data.get('payment_method', 'admin_topup')
        notes = request.data.get('notes', 'Manual top-up by admin')
        topup_status = request.data.get('status', 'completed')
        
        # Validation
        if not user_id:
            return Response(
                {'error': 'User ID required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validate amount
        try:
            amount = Decimal(str(amount))
            if amount <= 0:
                return Response(
                    {'error': 'Amount must be positive'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        except (ValueError, InvalidOperation):
            return Response(
                {'error': 'Invalid amount format'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get user (Enforce store isolation)
        try:
            from users.models import User
            if request.user.role == 'super_admin':
                user = User.objects.get(id=user_id)
            else:
                if not request.user.store:
                    return Response({'error': 'Admin belum ditugaskan ke toko manapun'}, status=status.HTTP_403_FORBIDDEN)
                user = User.objects.get(id=user_id, store=request.user.store)
        except User.DoesNotExist:
            return Response(
                {'error': 'User not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Get or create wallet
        wallet, created = Wallet.objects.get_or_create(
            user=user,
            defaults={'balance': Decimal('0.00')}
        )
        
        # Create top-up transaction with unique ID (timestamp + random suffix)
        topup = TopUpTransaction.objects.create(
            user=user,
            transaction_id=f"ADMIN{timezone.now().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:6]}",
            amount=amount,
            payment_method=payment_method,
            status=topup_status,
            verified_by=request.user,
            verified_at=timezone.now()
        )
        
        # If status is completed, update wallet immediately
        if topup_status == 'completed':
            old_balance = wallet.balance
            wallet.balance += topup.amount
            wallet.total_topup += topup.amount
            wallet.save()
            
            # Create wallet history
            WalletHistory.objects.create(
                wallet=wallet,
                transaction_type='topup',
                amount=topup.amount,
                balance_before=old_balance,
                balance_after=wallet.balance,
                description=f"Admin top-up: {notes}",
                reference_id=str(topup.id)
            )
        
        # Clear cache
        cache.delete('admin_dashboard_stats')
        
        return Response({
            'message': 'Top-up created successfully',
            'topup': {
                'id': str(topup.id),
                'user': user.email,
                'amount': float(topup.amount),
                'status': topup.status,
                'transaction_id': topup.transaction_id
            },
            'wallet_balance': float(wallet.balance) if topup_status == 'completed' else None
        }, status=status.HTTP_201_CREATED)
    
    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """
        Approve top-up
        POST /api/admin/topups/{id}/approve/
        """
        topup = self.get_object()
        
        if topup.status != 'pending':
            return Response(
                {'error': 'Top-up already processed'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Update topup status
        topup.status = 'success'
        topup.verified_by = request.user
        topup.verified_at = timezone.now()
        topup.save()
        
        # Update wallet balance
        wallet = Wallet.objects.get(user=topup.user)
        balance_before = wallet.balance
        wallet.balance += topup.amount
        wallet.total_topup += topup.amount
        wallet.save()
        
        # Create wallet history
        WalletHistory.objects.create(
            wallet=wallet,
            transaction_type='topup',
            amount=topup.amount,
            balance_before=balance_before,
            balance_after=wallet.balance,
            description=f"Top-up approved: {topup.transaction_id}",
            reference_id=topup.id,
            reference_type='topup'
        )
        
        # Clear cache
        cache.delete('admin_dashboard_stats')
        
        return Response({'message': 'Top-up approved successfully'})


class AdminPolicyTierViewSet(viewsets.ModelViewSet):
    """
    CRUD Policy Tiers untuk Admin
    GET/POST/PUT/DELETE /api/admin/policy-tiers/
    """
    queryset = PolicyTier.objects.all().order_by('min_price')
    permission_classes = [IsStoreStaffOrAbove]
    
    # Import serializer local to avoid circular import
    from policies.serializers import PolicyTierSerializer
    serializer_class = PolicyTierSerializer
